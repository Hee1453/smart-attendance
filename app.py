import uuid
from datetime import datetime, timedelta
import math
import psycopg2
import psycopg2.extras
import pandas as pd
import os
from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from authlib.integrations.flask_client import OAuth
import json

app = Flask(__name__)

# ==========================================
# ⚙️ ตั้งค่าระบบ และ Google OAuth (ดึงจาก Environment)
# ==========================================
app.secret_key = os.environ.get('SECRET_KEY', 'super_secret_key_change_this')

app.config['GOOGLE_CLIENT_ID'] = os.environ.get('GOOGLE_CLIENT_ID', '1055465619000-mi7kalvlqi6cuumuqholbqhm6bi5et7b.apps.googleusercontent.com')
app.config['GOOGLE_CLIENT_SECRET'] = os.environ.get('GOOGLE_CLIENT_SECRET', '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')

oauth = OAuth(app)
google = oauth.register(
    name='google',
    client_id=app.config['GOOGLE_CLIENT_ID'],
    client_secret=app.config['GOOGLE_CLIENT_SECRET'],
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={'scope': 'openid email profile'}
)

def get_thai_now():
    return datetime.utcnow() + timedelta(hours=7)

# 👑 ตั้งค่า Super Admin 
SUPER_ADMIN_EMAIL = os.environ.get('SUPER_ADMIN_EMAIL', 'your_real_email@gmail.com')

# URL สำหรับเชื่อมต่อ Database Neon
DATABASE_URL = os.environ.get('DATABASE_URL', 'postgresql://neondb_owner:npg_zmaLVEd9vt8C@ep-holy-breeze-a1p4sqrq-pooler.ap-southeast-1.aws.neon.tech/neondb?sslmode=require&channel_binding=require')

def get_db():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.DictCursor)
    conn.autocommit = True 
    return conn

def init_db():
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('CREATE TABLE IF NOT EXISTS sessions (id SERIAL PRIMARY KEY, subject_id TEXT, created_at TEXT, teacher_email TEXT)')
    
    cursor.execute("SELECT column_name FROM information_schema.columns WHERE table_name='sessions' AND column_name='teacher_email'")
    if not cursor.fetchone():
        cursor.execute("ALTER TABLE sessions ADD COLUMN teacher_email TEXT DEFAULT 'unknown'")
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS attendance (
            id SERIAL PRIMARY KEY, 
            session_id INTEGER, 
            student_id TEXT, 
            check_in_time TEXT, 
            distance TEXT, 
            email TEXT,
            name TEXT,
            picture TEXT,
            status TEXT,
            ip_address TEXT,
            device_info TEXT,
            FOREIGN KEY(session_id) REFERENCES sessions(id)
        )
    ''')
    
    cursor.execute('CREATE TABLE IF NOT EXISTS teachers (id SERIAL PRIMARY KEY, email TEXT UNIQUE)')
    
    cursor.close()
    conn.close()

init_db()

# ==========================================
# 🧠 ระบบ Memory แยกตามอีเมลอาจารย์ (Multi-Teacher)
# ==========================================
active_sessions = {}

def get_teacher_session(email):
    if email not in active_sessions:
        active_sessions[email] = {
            "is_active": False, "db_id": None, "subject_id": None, "teacher_lat": None, "teacher_long": None,
            "radius": 50, "time_limit": 15, "start_time": None, "current_qr_token": None, 
            "attendees": [], "roster": []
        }
    return active_sessions[email]

def haversine_distance(lat1, lon1, lat2, lon2):
    R = 6371 
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c * 1000 

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/login')
def login():
    redirect_uri = url_for('authorize', _external=True)
    return google.authorize_redirect(redirect_uri)

@app.route('/authorize')
def authorize():
    token = google.authorize_access_token()
    user_info = token.get('userinfo')
    session['user'] = user_info
    email = user_info['email']
    
    conn = get_db()
    cursor = conn.cursor()
    
    # เช็คว่าเป็นอาจารย์ไหม
    cursor.execute('SELECT email FROM teachers WHERE email = %s', (email,))
    is_teacher = cursor.fetchone()
    
    if is_teacher or email == SUPER_ADMIN_EMAIL:
        cursor.close(); conn.close()
        session['role'] = 'teacher'
        return redirect('/teacher') 
        
    # เช็คว่าใช้อีเมลมหาลัยไหม
    if not email.endswith('@rmutsb.ac.th'):
        cursor.close(); conn.close()
        session.clear() 
        return "⛔ เข้าสู่ระบบล้มเหลว: สำหรับนักศึกษา กรุณาใช้อีเมลของมหาวิทยาลัย (@rmutsb.ac.th) เท่านั้น", 403

    # ตัดรหัสนักศึกษา
    session['role'] = 'student'
    try:
        temp_id = email.split('@')[0]
        student_id = temp_id[:12]
    except:
        student_id = email[:12]
    session['student_id'] = student_id
    
    # 🌟 [เพิ่มใหม่] เช็คว่าเคยมีชื่อในระบบ (เคยเช็คชื่อ) แล้วหรือยัง
    cursor.execute('SELECT name FROM attendance WHERE student_id = %s AND name IS NOT NULL AND name != \'ไม่ระบุชื่อ\' LIMIT 1', (student_id,))
    existing_user = cursor.fetchone()
    cursor.close(); conn.close()
    
    # ถ้าเคยมีชื่อแล้ว ให้เซฟลง Session แล้วข้ามไปหน้า Student เลย ไม่ต้อง Setup อีก
    if existing_user:
        session['user']['name'] = existing_user['name']
        return redirect('/student')
        
    # ถ้าเป็นเด็กใหม่กิ๊ก ค่อยพาไปหน้า Setup
    return redirect('/setup_profile')

@app.route('/logout')
def logout():
    session.pop('user', None)
    session.pop('student_id', None)
    session.pop('role', None)
    return redirect('/')

@app.route('/student')
def student_page():
    user = session.get('user')
    if not user: return redirect('/login') 
    
    student_id = session.get('student_id')
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT attendance.*, sessions.subject_id, sessions.created_at as class_date
        FROM attendance
        JOIN sessions ON attendance.session_id = sessions.id
        WHERE attendance.student_id = %s
        ORDER BY sessions.created_at DESC
    ''', (student_id,))
    raw_history = cursor.fetchall()
    cursor.close()
    conn.close()

    TOTAL_CLASSES_PER_SUBJECT = 18
    subjects_data = {}

    for row in raw_history:
        subj = row['subject_id']
        if subj not in subjects_data:
            subjects_data[subj] = {
                "subject_id": subj,
                "attended": 0,
                "total": TOTAL_CLASSES_PER_SUBJECT,
                "percent": 0,
                "history": []
            }
        
        subjects_data[subj]["history"].append(row)
        subjects_data[subj]["attended"] += 1

    for subj in subjects_data:
        percent = (subjects_data[subj]["attended"] / TOTAL_CLASSES_PER_SUBJECT) * 100
        subjects_data[subj]["percent"] = int(percent)

    return render_template('student.html', user=user, student_id=student_id, subjects_data=subjects_data)

@app.route('/teacher')
def teacher_page():
    user = session.get('user')
    if not user: return redirect('/login')
        
    email = user.get('email')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT email FROM teachers WHERE email = %s', (email,))
    is_teacher = cursor.fetchone()
    cursor.close()
    conn.close()
        
    if not is_teacher and email != SUPER_ADMIN_EMAIL:
        return "⛔ ไม่อนุญาตให้เข้าถึง: หน้านี้สงวนไว้สำหรับอาจารย์ผู้สอนเท่านั้น", 403
        
    return render_template('teacher.html', user=user)

@app.route('/attendance_records')
def attendance_records():
    user = session.get('user')
    if not user: return redirect('/login')
    
    curr_sess = get_teacher_session(user['email'])
    sorted_attendees = sorted(curr_sess['attendees'], key=lambda x: (x['id'][-3:], x['id']))
    return render_template('attendance_records.html', attendees=sorted_attendees, subject=curr_sess.get('subject_id'), current_session=curr_sess)

@app.route('/history')
def history_page():
    user = session.get('user')
    if not user: return redirect('/login')
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM sessions WHERE teacher_email = %s ORDER BY id DESC', (user['email'],))
    sessions = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return render_template('history.html', sessions=sessions)

@app.route('/history/<int:session_id>')
def history_detail(session_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM sessions WHERE id = %s', (session_id,))
    session_data = cursor.fetchone()
    
    cursor.execute('''
        SELECT * FROM attendance 
        WHERE session_id = %s 
        ORDER BY RIGHT(student_id, 3) ASC, student_id ASC
    ''', (session_id,))
    students = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    if not session_data: return "ไม่พบข้อมูลวิชานี้", 404
    return render_template('history_detail.html', session=session_data, students=students)

@app.route('/export_history/<int:session_id>')
def export_history(session_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT subject_id, created_at FROM sessions WHERE id = %s', (session_id,))
    session_info = cursor.fetchone()
    
    cursor.execute('''
        SELECT student_id, name, check_in_time, distance, status 
        FROM attendance 
        WHERE session_id = %s 
        ORDER BY RIGHT(student_id, 3) ASC, student_id ASC
    ''', (session_id,))
    students = cursor.fetchall()
    cursor.close()
    conn.close()
    
    if not students: return "ไม่มีข้อมูลให้ Export"
    
    status_map = {'present': 'มาเรียน', 'late': 'มาสาย', 'leave': 'ลาป่วย/ลากิจ'}
    data_list = []
    for row in students:
        raw_status = row['status'] if 'status' in row.keys() else 'present'
        data_list.append({
            "รหัสนักศึกษา": row['student_id'],
            "ชื่อ-นามสกุล": row['name'] if 'name' in row.keys() and row['name'] else "ไม่ระบุ",
            "เวลาที่เช็คชื่อ": row['check_in_time'],
            "ระยะห่าง": row['distance'],
            "สถานะ": status_map.get(raw_status, raw_status)
        })
        
    df = pd.DataFrame(data_list)
    subject_name = session_info['subject_id'] if session_info else "Class"
    filename = f"History_{subject_name}_{session_id}.xlsx"
    
    from openpyxl.utils import get_column_letter
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']
        for idx, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 5
            col_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[col_letter].width = max_len
            
    return send_file(filename, as_attachment=True)

# API สำหรับเพิ่มชื่อแบบ Manual ย้อนหลัง (เมื่อคลาสปิดไปแล้ว)
@app.route('/api/history_add_student', methods=['POST'])
def history_add_student():
    if not session.get('user'): return jsonify({"status": "error", "message": "กรุณาล็อกอิน"}), 401
    
    data = request.json
    session_id = data.get('session_id')
    student_id = data.get('student_id')
    req_name = data.get('name', '').strip()
    status = data.get('status', 'present')
    time_str = get_thai_now().strftime("%H:%M:%S")
    
    conn = get_db()
    cursor = conn.cursor()
    
    # 1. เช็คว่ามีชื่อในคาบนี้หรือยัง
    cursor.execute('SELECT id FROM attendance WHERE session_id = %s AND student_id = %s', (session_id, student_id))
    if cursor.fetchone():
        cursor.close(); conn.close()
        return jsonify({"status": "error", "message": "นักศึกษาคนนี้มีชื่อในคลาสนี้แล้ว"})
        
    # 2. หาชื่อจากประวัติเก่า
    cursor.execute('SELECT name, picture FROM attendance WHERE student_id = %s LIMIT 1', (student_id,))
    student_info = cursor.fetchone()
    
    final_name = req_name if req_name else (student_info['name'] if student_info and student_info['name'] else 'เพิ่มโดยอาจารย์ (ย้อนหลัง)')
    picture = student_info['picture'] if student_info and student_info['picture'] else ''
    
    # 3. บันทึกลง DB
    cursor.execute('''
        INSERT INTO attendance (session_id, student_id, check_in_time, distance, email, name, picture, status, ip_address, device_info) 
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    ''', (session_id, student_id, time_str, 'Manual', '', final_name, picture, status, 'Manual', 'Manual'))
    
    conn.commit()
    cursor.close()
    conn.close()
    
    return jsonify({"status": "success", "message": "เพิ่มข้อมูลย้อนหลังเรียบร้อย"})

@app.route('/setup_profile')
def setup_profile_page():
    user = session.get('user')
    if not user: return redirect('/login')
    # ดึงค่าเก่าที่พิมพ์ไว้มาแสดงเผื่อกด "แก้ไข"
    fname = ""
    lname = ""
    if 'name' in user and " " in user['name']:
        parts = user['name'].split(" ", 1)
        fname = parts[0]
        lname = parts[1]
    return render_template('setup_profile.html', user=user, student_id=session.get('student_id'), fname=fname, lname=lname)

@app.route('/save_profile', methods=['POST'])
def save_profile():
    if 'user' not in session: return redirect('/login')
    fname = request.form.get('fname')
    lname = request.form.get('lname')
    full_name = f"{fname} {lname}"
    
    # อัปเดตใน Session
    user_info = session['user']
    user_info['name'] = full_name
    session['user'] = user_info
    
    # 🌟 [เพิ่มใหม่] แอบไปอัปเดตชื่อใน Database ย้อนหลังให้ด้วย
    student_id = session.get('student_id')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE attendance SET name = %s WHERE student_id = %s', (full_name, student_id))
    conn.commit()
    cursor.close()
    conn.close()
    
    return redirect('/student')

@app.route('/api/delete_session', methods=['POST'])
def delete_session():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM attendance WHERE session_id = %s', (data['id'],))
    cursor.execute('DELETE FROM sessions WHERE id = %s', (data['id'],))
    cursor.close()
    conn.close()
    return jsonify({"status": "success"})

@app.route('/api/edit_session', methods=['POST'])
def edit_session():
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('UPDATE sessions SET subject_id = %s WHERE id = %s', (data['new_name'], data['id']))
    cursor.close()
    conn.close()
    return jsonify({"status": "success"})

@app.route('/export_excel')
def export_live_excel():
    user = session.get('user')
    if not user: return "กรุณาล็อกอิน"
    curr_sess = get_teacher_session(user['email'])
    
    if not curr_sess['attendees']: return "ไม่มีข้อมูลให้ Export"
    
    sorted_attendees = sorted(curr_sess['attendees'], key=lambda x: (x['id'][-3:], x['id']))
    df = pd.DataFrame(sorted_attendees)
    
    subject_name = curr_sess.get('subject_id', 'Unknown')
    df.insert(0, 'subject_id', subject_name)
    
    columns_map = {'subject_id': 'วิชา', 'id': 'รหัสนักศึกษา', 'name': 'ชื่อ-สกุล', 'time': 'เวลาที่มา', 'dist': 'ระยะห่าง', 'status': 'สถานะ'}
    existing_cols = [c for c in columns_map.keys() if c in df.columns]
    df = df[existing_cols]
    df.rename(columns=columns_map, inplace=True)
    
    status_map = {'present': 'มาเรียน', 'late': 'มาสาย', 'leave': 'ลาป่วย/ลากิจ'}
    df['สถานะ'] = df['สถานะ'].map(lambda x: status_map.get(x, x))
    
    filename = f"Attendance_{subject_name}_{get_thai_now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    
    from openpyxl.utils import get_column_letter
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        worksheet = writer.sheets['Sheet1']
        for idx, col in enumerate(df.columns):
            max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 5
            col_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[col_letter].width = max_len
            
    return send_file(filename, as_attachment=True)

@app.route('/api/start_class', methods=['POST'])
def start_class():
    user = session.get('user')
    if not user: return jsonify({"status": "error", "message": "กรุณาล็อกอิน"})
    
    data = request.json
    conn = get_db()
    cursor = conn.cursor()
    now_thai = get_thai_now() 
    now_str = now_thai.strftime("%Y-%m-%d %H:%M:%S")
    teacher_email = user['email']
    
    cursor.execute('INSERT INTO sessions (subject_id, created_at, teacher_email) VALUES (%s, %s, %s) RETURNING id', (data['subject_id'], now_str, teacher_email))
    new_db_id = cursor.fetchone()[0]
    
    cursor.close()
    conn.close()
    
    raw_roster = data.get('roster', '')
    roster_list = [x.strip() for x in raw_roster.replace(',', '\n').split('\n') if x.strip()]
    
    curr_sess = get_teacher_session(teacher_email)
    curr_sess.update({
        "is_active": True, 
        "db_id": new_db_id, 
        "subject_id": data['subject_id'],
        "teacher_lat": float(data['lat']), 
        "teacher_long": float(data['lng']),
        "radius": int(data['radius']), 
        "time_limit": int(data['time_limit']),
        "start_time": now_thai,
        "current_qr_token": str(uuid.uuid4())[:8], 
        "roster": roster_list,
        "attendees": []
    })
    return jsonify({"status": "success", "message": "Class Started"})

@app.route('/api/update_qr_token', methods=['GET'])
def update_qr_token():
    user = session.get('user')
    if not user: return jsonify({"status": "expired"})
    
    curr_sess = get_teacher_session(user['email'])
    if not curr_sess['is_active']: return jsonify({"status": "expired"})
    
    elapsed = (get_thai_now() - curr_sess['start_time']).total_seconds() / 60
    if elapsed > curr_sess['time_limit']:
        curr_sess['is_active'] = False
        return jsonify({"status": "expired"})
        
    return jsonify({"qr_token": curr_sess['current_qr_token'], "time_left": curr_sess['time_limit'] - elapsed})

@app.route('/api/get_dashboard_data', methods=['GET'])
def get_dashboard_data():
    user = session.get('user')
    if not user: return jsonify({"status": "error"})
    
    curr_sess = get_teacher_session(user['email'])
    sorted_attendees = sorted(curr_sess['attendees'], key=lambda x: (x['id'][-3:], x['id']))
    
    present_ids = [s['id'] for s in sorted_attendees]
    absent_list = [uid for uid in curr_sess['roster'] if uid not in present_ids]
    return jsonify({
        "attendees": sorted_attendees,
        "absent_list": absent_list,
        "total_students": len(curr_sess['roster'])
    })

@app.route('/api/check_in', methods=['POST'])
def check_in():
    user = session.get('user')
    student_id = session.get('student_id')
    
    if not user or not student_id: return jsonify({"status": "error", "message": "กรุณาล็อกอินใหม่"})
    data = request.json
    token = data.get('qr_token')
    
    target_session = None
    for s in active_sessions.values():
        if s['is_active'] and s['current_qr_token'] == token:
            target_session = s
            break
            
    if not target_session: 
        return jsonify({"status": "error", "message": "QR Code ไม่ถูกต้อง/หมดอายุ หรือคลาสถูกปิดไปแล้ว"})

    dist = haversine_distance(target_session['teacher_lat'], target_session['teacher_long'], float(data['lat']), float(data['lng']))
    if dist > target_session['radius']: return jsonify({"status": "error", "message": f"อยู่นอกพื้นที่ ({dist:.0f} เมตร)"})

    if any(s['id'] == student_id for s in target_session['attendees']): return jsonify({"status": "error", "message": "คุณเช็คชื่อไปแล้ว"})

    if request.headers.getlist("X-Forwarded-For"):
        client_ip = request.headers.getlist("X-Forwarded-For")[0].split(',')[0].strip()
    else:
        client_ip = request.remote_addr

    user_agent = request.headers.get('User-Agent')

    # [ระบบป้องกันการสแกนซ้ำ] 
    for s in target_session['attendees']:
        saved_ip = s.get('ip')
        saved_ua = s.get('ua')
        if saved_ip == client_ip and saved_ua == user_agent:
             return jsonify({
                 "status": "error", 
                 "message": "⛔ ไม่สามารถเช็คชื่อได้: ตรวจพบการใช้อุปกรณ์ซ้ำกับรหัส " + s['id']
             })

    now_thai = get_thai_now()
    elapsed_minutes = (now_thai - target_session['start_time']).total_seconds() / 60
    status = "late" if elapsed_minutes > 15 else "present"
    time_str = now_thai.strftime("%H:%M:%S")
    
    student_record = {
        "id": student_id, "time": time_str, "dist": f"{dist:.0f}m",
        "name": user.get('name', 'ไม่ระบุชื่อ'), "picture": user.get('picture', ''), 
        "status": status,
        "ip": client_ip,      
        "ua": user_agent      
    }
    target_session['attendees'].append(student_record)
    target_session['current_qr_token'] = str(uuid.uuid4())[:8]

    if target_session['db_id']:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO attendance (session_id, student_id, check_in_time, distance, email, name, picture, status, ip_address, device_info) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            target_session['db_id'], student_id, time_str, f"{dist:.0f}m", 
            user.get('email', ''), user.get('name', ''), user.get('picture', ''), status,
            client_ip, user_agent
        ))
        cursor.close()
        conn.close()

    return jsonify({"status": "checked_in"})

# ==========================================
# ส่วนของ ADMIN
# ==========================================
ADMIN_PASSWORD = "1234"

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        if request.form.get('password') == ADMIN_PASSWORD:
            session['is_admin'] = True
            return redirect('/admin')
        else: return render_template('admin_login.html', error="รหัสผ่านไม่ถูกต้อง")
    return render_template('admin_login.html')

@app.route('/admin/logout')
def admin_logout():
    session.pop('is_admin', None)
    return redirect('/admin/login')

@app.route('/admin')
def admin_dashboard():
    if not session.get('is_admin'): return redirect('/admin/login')
    
    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute('SELECT COUNT(*) FROM sessions')
    total_sessions = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM attendance')
    total_checkins = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(DISTINCT student_id) FROM attendance')
    unique_students = cursor.fetchone()[0]
    
    stats = {
        'total_sessions': total_sessions,
        'total_checkins': total_checkins,
        'unique_students': unique_students
    }
    
    cursor.execute('SELECT * FROM sessions ORDER BY created_at DESC')
    sessions = cursor.fetchall()
    
    risk_students = []
    if total_classes := stats['total_sessions']:
        cursor.execute('''
            SELECT student_id, name, COUNT(*) as attended_count
            FROM attendance
            GROUP BY student_id, name
        ''')
        student_stats = cursor.fetchall()
        
        for s in student_stats:
            percent = (s['attended_count'] / total_classes) * 100
            if percent < 80:
                risk_students.append({
                    'id': s['student_id'],
                    'name': s['name'],
                    'attended': s['attended_count'],
                    'total': total_classes,
                    'percent': int(percent)
                })

    cursor.execute('''
        SELECT substr(created_at, 1, 10) as date, COUNT(*) as count 
        FROM sessions 
        GROUP BY date 
        ORDER BY date DESC LIMIT 7
    ''')
    graph_data = cursor.fetchall()
    
    cursor.execute('''
        SELECT sessions.subject_id, MAX(attendance.check_in_time) as created_at, attendance.ip_address, COUNT(DISTINCT attendance.student_id) as dup_count
        FROM attendance
        JOIN sessions ON attendance.session_id = sessions.id
        WHERE attendance.ip_address != 'Manual' 
        GROUP BY attendance.session_id, attendance.ip_address, sessions.subject_id
        HAVING COUNT(DISTINCT attendance.student_id) > 1
        ORDER BY created_at DESC
    ''')
    cheating_logs = cursor.fetchall()
    
    cursor.execute('SELECT * FROM teachers ORDER BY id DESC')
    teachers_list = cursor.fetchall()

    cursor.close()
    conn.close()
    
    chart_labels = [row['date'] for row in graph_data][::-1]
    chart_values = [row['count'] for row in graph_data][::-1]

    return render_template('admin.html', 
                           stats=stats, 
                           sessions=sessions, 
                           risk_students=risk_students,
                           cheating_logs=cheating_logs,
                           teachers_list=teachers_list,
                           chart_labels=json.dumps(chart_labels),
                           chart_values=json.dumps(chart_values))

@app.route('/api/admin/reset_database', methods=['POST'])
def admin_reset_db():
    if not session.get('is_admin'): return jsonify({"status": "error", "message": "Unauthorized"}), 403
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('TRUNCATE TABLE attendance, sessions RESTART IDENTITY CASCADE')
        cursor.close()
        conn.close()
        return jsonify({"status": "success", "message": "ล้างข้อมูลเรียบร้อยแล้ว"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

@app.route('/api/admin/add_teacher', methods=['POST'])
def add_teacher():
    if not session.get('is_admin'): return jsonify({"status": "error"}), 403
    email = request.json.get('email', '').strip()
    if not email: return jsonify({"status": "error", "message": "กรุณากรอกอีเมล"})
    
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('INSERT INTO teachers (email) VALUES (%s)', (email,))
        conn.close()
        return jsonify({"status": "success"})
    except psycopg2.IntegrityError:
        return jsonify({"status": "error", "message": "อีเมลนี้มีอยู่ในระบบแล้ว"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)})

@app.route('/api/admin/delete_teacher', methods=['POST'])
def delete_teacher():
    if not session.get('is_admin'): return jsonify({"status": "error"}), 403
    teacher_id = request.json.get('id')
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM teachers WHERE id = %s', (teacher_id,))
    conn.close()
    return jsonify({"status": "success"})

@app.route('/api/stop_class', methods=['POST'])
def stop_class():
    user = session.get('user')
    if not user: return jsonify({"status": "error", "message": "กรุณาล็อกอิน"})
    
    curr_sess = get_teacher_session(user['email'])
    if curr_sess['is_active']:
        curr_sess['is_active'] = False
        return jsonify({"status": "success", "message": "ปิดคลาสเรียบร้อย"})
    return jsonify({"status": "error", "message": "ไม่มีคลาสที่เปิดอยู่"})

@app.route('/api/manual_checkin', methods=['POST'])
def manual_checkin():
    user = session.get('user')
    if not user: return jsonify({"status": "error", "message": "กรุณาล็อกอิน"})
    
    curr_sess = get_teacher_session(user['email'])
    if not curr_sess['is_active']:
        return jsonify({"status": "error", "message": "ไม่มีคลาสที่กำลังเปิดอยู่ กรุณาเปิดคลาสก่อน"})
        
    data = request.json
    student_id = data.get('id')
    req_name = data.get('name', '').strip() 
    time_str = data.get('time')
    dist_str = data.get('dist', 'Manual')
    req_status = data.get('status', 'present')
    
    if any(s['id'] == student_id for s in curr_sess['attendees']):
        return jsonify({"status": "error", "message": "นักศึกษาคนนี้มีชื่อในระบบแล้ว"})
        
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT name, picture FROM attendance WHERE student_id = %s LIMIT 1', (student_id,))
    student_info = cursor.fetchone()
    
    if req_name:
        final_name = req_name
    elif student_info and student_info['name']:
        final_name = student_info['name'] 
    else:
        final_name = 'ไม่ระบุชื่อ' 
        
    picture = student_info['picture'] if student_info and student_info['picture'] else ''
    
    student_record = {
        "id": student_id, "time": time_str, "dist": dist_str,
        "name": final_name, "picture": picture, "status": req_status,
        "ip": "Manual", "ua": "Manual"
    }
    curr_sess['attendees'].append(student_record)
    
    if curr_sess['db_id']:
        cursor.execute('''
            INSERT INTO attendance (session_id, student_id, check_in_time, distance, email, name, picture, status, ip_address, device_info) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (
            curr_sess['db_id'], student_id, time_str, dist_str, 
            '', final_name, picture, req_status, 'Manual', 'Manual'
        ))
        conn.commit()
        
    cursor.close()
    conn.close()
    
    return jsonify({"status": "success"})

if __name__ == '__main__':
    os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
    app.run(host='0.0.0.0', port=5000)